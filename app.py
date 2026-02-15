import openpyxl

book = openpyxl.Workbook()
print(book.sheetnames)

book.create_sheet('invoices')
invoices_page = book['invoices']

invoices_page.append(['date','product', 'produckt code', 'price'])
invoices_page.append(['22/02', 'yellow shirt', '179800', '55,99'])
invoices_page.append(['22/02', 'yellow t-shirt', '189800','55,99'])
invoices_page.append(['22/02', 'black shirt', '179800', '55,99'])
invoices_page.append(['22/02', 'pink t-shirt', '189800','55,99'])
invoices_page.append(['22/02', 'pink shirt', '179800', '55,99'])
invoices_page.append(['22/02', 'black t-shirt', '189800','55,99'])

book.save('Invoice.xlsx')